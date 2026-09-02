"""Excel (.xlsx) table for 批量导入. Same columns as CSV."""

from __future__ import annotations

import csv
import io

from backend.table_csv import CSV_FIELDS, _cell, parse_table_csv


def _openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise ValueError("缺少 Excel 支持，请安装 openpyxl") from exc
    return Workbook, load_workbook


def export_table_xlsx(items: list[dict], file_name: str = "") -> bytes:
    Workbook, _load = _openpyxl()
    book = Workbook()
    sheet = book.active
    sheet.title = "图译"
    sheet.append(list(CSV_FIELDS))
    for item in items or []:
        sheet.append(
            [
                file_name or _cell(item.get("file")),
                _cell(item.get("handle")),
                _cell(item.get("field") or "text"),
                _cell(item.get("source")),
                _cell(item.get("target")),
                _cell(item.get("layer")),
                _cell(item.get("type")),
            ]
        )
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def parse_table_xlsx(payload: bytes) -> list[dict]:
    _Workbook, load_workbook = _openpyxl()
    if not payload:
        return []
    book = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = book.active
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    for row in sheet.iter_rows(values_only=True):
        cells = ["" if cell is None else str(cell) for cell in row]
        if any(str(cell).strip() for cell in cells):
            writer.writerow(cells)
    return parse_table_csv(buffer.getvalue())


def parse_table_payload(*, csv: str = "", xlsx: bytes | None = None) -> list[dict]:
    if xlsx:
        return parse_table_xlsx(xlsx)
    return parse_table_csv(csv)
